#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
激光切割时间统计
从 PDF 文件名提取厚度，从 PDF 内容提取切割时间，生成完整艾威模板格式 Excel

版本：v5.0
更新：
- 删除空行后自动更新 F/H/L 列公式
- D 列留空，E 列填总和（两位小数）
- 合计公式放在备注行，保留两位小数，居中
- 备注行添加边框
- 完整艾威模板格式
"""

import os
import re
import tempfile
import py7zr
import pdfplumber
import openpyxl
import shutil
import glob
from typing import Dict, List, Optional
from collections import defaultdict
from openpyxl.styles import Border, Side, Alignment


class CuttingTimeCalculator:
    """激光切割时间计算器（艾威模板格式）"""
    
    def __init__(self, template_path: str = None):
        if not template_path:
            # 使用第一个可用的模板文件
            templates = glob.glob('/home/admin/.openclaw/media/inbound/艾威时间模板*.xlsx')
            if templates:
                self.template_path = templates[0]
            else:
                raise FileNotFoundError("找不到艾威时间模板文件")
        else:
            self.template_path = template_path
        
        # 厚度映射（PDF 厚度 → 模板行号）
        self.thickness_to_row = {
            1: 19, 1.5: 20, 2: 21, 2.5: 22, 3: 23,
            4: 24, 5: 25, 6: 26, 8: 27, 10: 28,
            12: 29, 14: 30, 16: 31
        }
    
    def get_thickness_from_filename(self, pdf_filename: str) -> Optional[float]:
        """从 PDF 文件名提取厚度"""
        match = re.search(r'(\d+\.?\d*)(Mn|mm|NM|nm)?', pdf_filename)
        if match:
            return float(match.group(1))
        return None
    
    def extract_cutting_time(self, pdf_path: str) -> List[float]:
        """从 PDF 提取切割时间（只提取 Job data 中的 Cutting time）"""
        times = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # 只处理第一页
                page = pdf.pages[0]
                text = page.extract_text() or ''
                lines = text.split('\n')
                
                in_job_data = False
                for line in lines:
                    if 'Job data' in line:
                        in_job_data = True
                        continue
                    
                    if in_job_data:
                        # 跳过表头行
                        if 'Job code' in line and 'Material' in line and 'Machine' in line:
                            continue
                        
                        # 查找实际数据行（Job code 行，包含 kg 和 min）
                        if 'kg' in line and 'min' in line and not line.strip().startswith('Job code'):
                            # 提取 Cutting time（在 kg 后面）
                            time_match = re.search(r'(\d+\.?\d*)\s*kg\s+(\d+\.?\d*)\s*min', line)
                            if time_match:
                                time_val = float(time_match.group(2))
                                times.append(time_val)
                            break
                        
                        # 遇到 Material data 或 Plan data 停止
                        if 'Material data' in line or 'Plan data' in line:
                            break
                    
        except Exception as e:
            print(f"PDF 处理失败 {pdf_path}: {e}")
        
        return times
    
    def process_zip(self, zip_path: str, output_path: str = None):
        """处理 ZIP/7Z 压缩包"""
        if not output_path:
            base_name = os.path.splitext(os.path.basename(zip_path))[0]
            output_path = os.path.join(os.path.dirname(zip_path), f'{base_name}_Cutting_time_statistics.xlsx')
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # 解压
            print(f"解压文件：{zip_path}")
            with py7zr.SevenZipFile(zip_path, mode='r') as z:
                z.extractall(tmpdir)
            
            # 查找所有 PDF
            pdf_files = []
            for root, dirs, files in os.walk(tmpdir):
                for f in files:
                    if f.lower().endswith('.pdf'):
                        pdf_files.append(os.path.join(root, f))
            
            # 按厚度排序
            def sort_key(path):
                basename = os.path.basename(path)
                thickness = self.get_thickness_from_filename(basename) or 999
                return (thickness, 0 if 'mm' in basename else 1)
            
            pdf_files.sort(key=sort_key)
            
            print(f"找到 {len(pdf_files)} 个 PDF 文件\n")
            
            # 按厚度分组
            time_by_thickness: Dict[float, List] = defaultdict(list)
            
            for pdf_path in pdf_files:
                basename = os.path.basename(pdf_path)
                pdf_thickness = self.get_thickness_from_filename(basename)
                
                if pdf_thickness:
                    times = self.extract_cutting_time(pdf_path)
                    
                    if times:
                        for t in times:
                            time_by_thickness[pdf_thickness].append({
                                'file': basename,
                                'time': t
                            })
                        total = sum(times)
                        print(f"✓ {basename}: {len(times)} 个时间，合计 {total:.2f}min")
                    else:
                        print(f"✗ {basename}: 未提取到切割时间")
            
            # 复制模板并填充
            self._copy_template_and_fill(time_by_thickness, output_path)
            print(f"\n✅ 已保存：{output_path}")
            
            return output_path
    
    def _copy_template_and_fill(self, time_by_thickness: Dict[float, List], output_path: str):
        """复制模板文件并填充数据，删除没有数据的行，更新公式"""
        # 复制到临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        shutil.copy2(self.template_path, tmp_path)
        
        # 打开模板
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
        ws = wb.active
        
        # 预先删除原 32 行（备注和合计），后面会添加到新位置
        ws.delete_rows(32, 1)
        
        # 填充数据
        sorted_thickness = sorted(time_by_thickness.keys())
        filled_rows = []  # 记录有数据的行
        
        for thickness in sorted_thickness:
            entries = time_by_thickness[thickness]
            total_time = sum(e['time'] for e in entries)
            
            # 找到模板行
            template_row = self.thickness_to_row.get(thickness)
            
            if template_row:
                row = template_row
                
                # D 列：留空（不填单个 PDF 时间）
                # E 列：直接填时间总和（保留两位小数）
                ws.cell(row=row, column=5, value=round(total_time, 2))
                filled_rows.append(row)
                
                print(f"  填充 {thickness}mm → 行{row}, {len(entries)} 个 PDF, 合计={total_time:.2f}min")
        
        # 不删除行，直接清除所有数据行（19-31 行）的 B-L 列
        for row in range(19, 32):
            for col in range(2, 13):  # B-L 列
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell._value = None
        
        # 保存一次确保清除生效
        wb.save(tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
        ws = wb.active
        
        # 填充有数据的厚度（连续填充到 19, 20, 21...）
        current_row = 19
        for thickness in sorted_thickness:
            total_time = sum(e['time'] for e in time_by_thickness[thickness])
            ws.cell(row=current_row, column=2, value=thickness)  # B 列厚度
            ws.cell(row=current_row, column=5, value=round(total_time, 2))  # E 列时间
            ws.cell(row=current_row, column=6, value=f'=E{current_row}+(E{current_row}*F$18)')
            ws.cell(row=current_row, column=7, value=6.33)
            ws.cell(row=current_row, column=8, value=f'=G{current_row}*F{current_row}')
            ws.cell(row=current_row, column=12, value=f'=H{current_row}+I{current_row}+J{current_row}+K{current_row}')
            print(f"  填充 {thickness}mm → 行{current_row}, 合计={total_time:.2f}min")
            current_row += 1
        
        last_data_row = current_row - 1
        remark_row = last_data_row + 1
        
        # 设置备注行
        ws.cell(row=remark_row, column=1, value='备注：')
        
        # 合计公式放在备注行
        if last_data_row >= 19:
            ws.cell(row=remark_row, column=12, value=f'=SUM(L19:L{last_data_row})')
            print(f"  更新合计公式：L{remark_row} =SUM(L19:L{last_data_row})")
        
        # 设置备注行边框和格式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col in range(1, 13):
            cell = ws.cell(row=remark_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # L 列（合计）：居中，保留两位小数
        l_cell = ws.cell(row=remark_row, column=12)
        l_cell.alignment = Alignment(horizontal='center', vertical='center')
        l_cell.number_format = '0.00'
        
        # 删除备注行之后的所有行
        while ws.max_row > remark_row:
            ws.delete_rows(remark_row + 1, 1)
        
        # 保存并重新加载，确保 max_row 正确
        wb.save(tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
        ws = wb.active
        
        wb.save(tmp_path)
        # 移动到目标位置
        shutil.move(tmp_path, output_path)


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("用法：python calculator.py <压缩包路径> [输出文件]")
        sys.exit(1)
    
    calculator = CuttingTimeCalculator()
    zip_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    calculator.process_zip(zip_path, output_path)
